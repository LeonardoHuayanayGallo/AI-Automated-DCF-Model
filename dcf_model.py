"""
dcf_model.py
Builds a fully formula-driven 5-year DCF model.
LLM assumptions populate INPUT cells.
All projection rows use live Excel formulas referencing those inputs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Color palette ──────────────────────────────────────────────────────────
BLUE_HEADER = "1F3864"
LIGHT_BLUE  = "D9E8F5"
GOLD        = "F2C94C"
GREEN       = "E2EFDA"
WHITE       = "FFFFFF"
DARK_GRAY   = "2C2C2C"
INPUT_YELLOW = "FFF2CC"

def style_header(cell, text, bg=BLUE_HEADER, size=11):
    cell.value = text
    cell.font  = Font(bold=True, color=WHITE, size=size)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_input(cell, value, fmt="0.00%"):
    cell.value = value
    cell.number_format = fmt
    cell.fill = PatternFill("solid", fgColor=INPUT_YELLOW)
    cell.font = Font(bold=True, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center")

def style_label(cell, text, bold=False):
    cell.value = text
    cell.font  = Font(bold=bold, color=DARK_GRAY, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_formula(cell, formula, fmt='$#,##0.0', bold=False, bg=None):
    cell.value = formula
    cell.number_format = fmt
    cell.font = Font(bold=bold, color=DARK_GRAY, size=10)
    cell.alignment = Alignment(horizontal="right")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def build_dcf(
    ticker: str,
    base_revenue: float,
    assumptions: dict,
    wacc: float = 0.10,
    terminal_growth: float = 0.025,
    output_dir: str = "output"
) -> str:

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{ticker}_DCF_Model.xlsx")
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1 — DCF MODEL
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "DCF Model"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18  # Base / Input
    for col in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    # Row heights
    for row in range(1, 50):
        ws.row_dimensions[row].height = 18

    # ── TITLE ──────────────────────────────────────────────────────────────
    ws["B2"] = f"{ticker} — AI-Generated DCF Valuation Model"
    ws["B2"].font = Font(bold=True, size=14, color=DARK_GRAY)
    ws["B3"] = "Source: SEC EDGAR 10-K  |  LLM: Meta Llama 3.3  |  Built with Python + OpenPyXL"
    ws["B3"].font = Font(italic=True, size=9, color="888888")

    # ── SECTION: ASSUMPTIONS (INPUT CELLS) ────────────────────────────────
    ws["B5"] = "MODEL ASSUMPTIONS (editable — model updates automatically)"
    ws["B5"].font = Font(bold=True, size=11, color=WHITE)
    ws["B5"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws.merge_cells("B5:H5")
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center")

    # Headers for assumption table
    for col, label in zip(["B","C"], ["Input Parameter", "Value"]):
        style_header(ws[f"{col}6"], label, bg="2E4057", size=10)

    # ── Input rows (these are the cells the formulas will reference) ──────
    # Row 7:  Base Revenue
    # Row 8:  Revenue Growth Y1
    # Row 9:  Revenue Growth Y2
    # Row 10: Revenue Growth Y3
    # Row 11: EBIT Margin
    # Row 12: Capex % Revenue
    # Row 13: Tax Rate
    # Row 14: D&A % Revenue (standard assumption)
    # Row 15: Change in NWC % Revenue (standard assumption)
    # Row 16: WACC
    # Row 17: Terminal Growth Rate

    input_params = [
        ("Base Revenue ($M) — LTM Actual",    base_revenue,                               "$#,##0.0"),
        ("Revenue Growth — Year 1 (LLM)",     assumptions["revenue_growth_y1"] / 100,     "0.0%"),
        ("Revenue Growth — Year 2 (LLM)",     assumptions["revenue_growth_y2"] / 100,     "0.0%"),
        ("Revenue Growth — Year 3 (LLM)",     assumptions["revenue_growth_y3"] / 100,     "0.0%"),
        ("EBIT Margin (LLM)",                 assumptions["ebit_margin"] / 100,           "0.0%"),
        ("Capex as % of Revenue (LLM)",       assumptions["capex_pct_revenue"] / 100,     "0.0%"),
        ("Effective Tax Rate (LLM)",          assumptions["tax_rate"] / 100,              "0.0%"),
        ("D&A as % of Revenue (assumed)",     0.04,                                       "0.0%"),
        ("Change in NWC as % Revenue",        0.02,                                       "0.0%"),
        ("WACC",                              wacc,                                       "0.0%"),
        ("Terminal Growth Rate",              terminal_growth,                            "0.0%"),
    ]

    # Named cell references for documentation
    # C7=Base Rev, C8=GY1, C9=GY2, C10=GY3, C11=EBIT margin
    # C12=Capex%, C13=Tax, C14=D&A%, C15=NWC%, C16=WACC, C17=TGR

    for i, (label, value, fmt) in enumerate(input_params):
        row = 7 + i
        style_label(ws[f"B{row}"], label)
        style_input(ws[f"C{row}"], value, fmt)
        ws[f"B{row}"].border = thin_border()
        ws[f"C{row}"].border = thin_border()

    # LLM Rationale (right side of assumption table)
    ws["E7"]  = "LLM Rationale"
    ws["E7"].font = Font(bold=True, color=WHITE, size=10)
    ws["E7"].fill = PatternFill("solid", fgColor="2E4057")
    ws.merge_cells("E7:H7")

    rationale_items = [
        assumptions["rationale"]["revenue_growth"],
        assumptions["rationale"]["revenue_growth"],
        assumptions["rationale"]["revenue_growth"],
        assumptions["rationale"]["ebit_margin"],
        assumptions["rationale"]["capex"],
        assumptions["rationale"]["tax_rate"],
    ]
    for i, text in enumerate(rationale_items):
        row = 8 + i
        ws[f"E{row}"] = text
        ws[f"E{row}"].font = Font(italic=True, size=9, color="555555")
        ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"E{row}:H{row}")

    # ── SECTION: PROJECTIONS (FORMULA-DRIVEN) ─────────────────────────────
    proj_start = 20

    ws[f"B{proj_start-1}"] = "5-YEAR FREE CASH FLOW PROJECTIONS"
    ws[f"B{proj_start-1}"].font = Font(bold=True, size=11, color=WHITE)
    ws[f"B{proj_start-1}"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws.merge_cells(f"B{proj_start-1}:H{proj_start-1}")
    ws[f"B{proj_start-1}"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers
    headers = ["Metric ($M)", "Base Year"] + [f"Year {i}" for i in range(1, 6)]
    for col_idx, h in enumerate(headers):
        col_letter = get_column_letter(col_idx + 2)
        style_header(ws[f"{col_letter}{proj_start}"], h, size=10)

    # ── Revenue row ────────────────────────────────────────────────────────
    # Year columns: D=Y1, E=Y2, F=Y3, G=Y4, H=Y5
    # Base year: C (hardcoded from input C7)
    r = proj_start + 1
    style_label(ws[f"B{r}"], "Revenue", bold=True)
    style_formula(ws[f"C{r}"], "=$C$7",         fmt="$#,##0.0", bg=LIGHT_BLUE)  # Base
    style_formula(ws[f"D{r}"], "=C{r}*(1+$C$8)".format(r=r), fmt="$#,##0.0", bg=LIGHT_BLUE)
    style_formula(ws[f"E{r}"], "=D{r}*(1+$C$9)".format(r=r), fmt="$#,##0.0", bg=LIGHT_BLUE)
    style_formula(ws[f"F{r}"], "=E{r}*(1+$C$10)".format(r=r), fmt="$#,##0.0", bg=LIGHT_BLUE)
    style_formula(ws[f"G{r}"], "=F{r}*(1+$C$10*0.85)".format(r=r), fmt="$#,##0.0", bg=LIGHT_BLUE)
    style_formula(ws[f"H{r}"], "=G{r}*(1+$C$10*0.70)".format(r=r), fmt="$#,##0.0", bg=LIGHT_BLUE)

    rev_row = r  # save for reference in later formulas

    # ── EBIT row ───────────────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "EBIT")
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{rev_row}*$C$11", fmt="$#,##0.0")

    ebit_row = r

    # ── NOPAT row ──────────────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "NOPAT  [EBIT × (1 − Tax Rate)]")
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{ebit_row}*(1-$C$13)", fmt="$#,##0.0")

    nopat_row = r

    # ── D&A row ────────────────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "(+) D&A")
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{rev_row}*$C$14", fmt="$#,##0.0")

    da_row = r

    # ── Capex row ──────────────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "(−) Capex")
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{rev_row}*$C$12", fmt="$#,##0.0")

    capex_row = r

    # ── ΔNWC row ───────────────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "(−) Change in NWC")
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{rev_row}*$C$15", fmt="$#,##0.0")

    nwc_row = r

    # ── Free Cash Flow row ─────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "Free Cash Flow to Firm (FCFF)", bold=True)
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{nopat_row}+{col}{da_row}-{col}{capex_row}-{col}{nwc_row}",
            fmt="$#,##0.0", bold=True, bg=LIGHT_BLUE)

    fcf_row = r

    # ── Discount Factor row ────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "Discount Factor  [1/(1+WACC)^t]")
    ws[f"C{r}"].value = "—"
    for t, col in enumerate(["D", "E", "F", "G", "H"], start=1):
        style_formula(ws[f"{col}{r}"],
            f"=1/(1+$C$16)^{t}", fmt="0.0000")

    df_row = r

    # ── PV of FCF row ──────────────────────────────────────────────────────
    r += 1
    style_label(ws[f"B{r}"], "PV of FCFF")
    ws[f"C{r}"].value = "—"
    for col in ["D", "E", "F", "G", "H"]:
        style_formula(ws[f"{col}{r}"],
            f"={col}{fcf_row}*{col}{df_row}", fmt="$#,##0.0", bg=GREEN)

    pv_row = r

    # ── VALUATION SUMMARY ──────────────────────────────────────────────────
    r += 2
    ws[f"B{r}"] = "VALUATION SUMMARY"
    ws[f"B{r}"].font = Font(bold=True, size=11, color=WHITE)
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")

    r += 1
    sum_pv_row = r
    style_label(ws[f"B{r}"], "Sum of PV(FCFF) — Years 1-5")
    style_formula(ws[f"D{r}"],
        f"=SUM(D{pv_row}:H{pv_row})", fmt="$#,##0.0")

    r += 1
    tv_row = r
    style_label(ws[f"B{r}"], "Terminal Value  [FCFF_5 × (1+TGR) / (WACC−TGR)]")
    style_formula(ws[f"D{r}"],
        f"=H{fcf_row}*(1+$C$17)/($C$16-$C$17)", fmt="$#,##0.0")

    r += 1
    pv_tv_row = r
    style_label(ws[f"B{r}"], "PV of Terminal Value")
    style_formula(ws[f"D{r}"],
        f"=D{tv_row}*H{df_row}", fmt="$#,##0.0")

    r += 1
    ev_row = r
    style_label(ws[f"B{r}"], "Enterprise Value", bold=True)
    ev_cell = ws[f"D{r}"]
    style_formula(ev_cell,
        f"=D{sum_pv_row}+D{pv_tv_row}",
        fmt="$#,##0.0", bold=True, bg=GOLD)
    ev_cell.font = Font(bold=True, size=12, color=DARK_GRAY)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2 — SENSITIVITY ANALYSIS (formula-driven)
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Sensitivity Analysis")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions["A"].width = 2
    ws2.column_dimensions["B"].width = 16
    for col in ["C","D","E","F","G"]:
        ws2.column_dimensions[col].width = 18

    ws2["B2"] = "Enterprise Value Sensitivity — WACC vs. Terminal Growth Rate"
    ws2["B2"].font = Font(bold=True, size=13, color=DARK_GRAY)
    ws2["B3"] = "All values in $M  |  Base case highlighted in gold  |  Formulas reference DCF Model sheet"
    ws2["B3"].font = Font(italic=True, size=9, color="888888")

    wacc_range = [wacc - 0.02, wacc - 0.01, wacc, wacc + 0.01, wacc + 0.02]
    tgr_range  = [0.015, 0.020, 0.025, 0.030, 0.035]

    # Axis labels
    ws2["B5"] = "TGR ↓  /  WACC →"
    ws2["B5"].font = Font(bold=True, color=WHITE)
    ws2["B5"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws2["B5"].alignment = Alignment(horizontal="center")

    for col_i, w in enumerate(wacc_range):
        col_letter = get_column_letter(col_i + 3)
        cell = ws2[f"{col_letter}5"]
        cell.value = w
        cell.number_format = "0.0%"
        cell.font  = Font(bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=BLUE_HEADER)
        cell.alignment = Alignment(horizontal="center")

    for row_i, tgr in enumerate(tgr_range):
        r2 = 6 + row_i
        # Row label
        label_cell = ws2[f"B{r2}"]
        label_cell.value = tgr
        label_cell.number_format = "0.0%"
        label_cell.font = Font(bold=True, color=WHITE)
        label_cell.fill = PatternFill("solid", fgColor=BLUE_HEADER)
        label_cell.alignment = Alignment(horizontal="center")

        for col_i, w in enumerate(wacc_range):
            col_letter = get_column_letter(col_i + 3)
            cell = ws2[f"{col_letter}{r2}"]

            # Formula: recalculate EV using different WACC and TGR
            # Sum of PV(FCF) changes with WACC, Terminal Value changes with both
            # We reference the FCF values from sheet 1 and recalculate inline
            w_val  = round(w, 4)
            tgr_val = round(tgr, 4)

            # Build the formula using the FCF row from sheet 1
            fcf_cells = [f"'DCF Model'!{col}{fcf_row}"
                         for col in ["D","E","F","G","H"]]

            pv_parts = "+".join(
                [f"{fcf_cells[t]}/(1+{w_val})^{t+1}" for t in range(5)]
            )
            tv_formula = (f"'DCF Model'!H{fcf_row}*(1+{tgr_val})"
                          f"/({w_val}-{tgr_val})"
                          f"/(1+{w_val})^5")

            cell.value = f"={pv_parts}+{tv_formula}"
            cell.number_format = "$#,##0.0"
            cell.alignment = Alignment(horizontal="right")

            # Highlight base case
            if col_i == 2 and row_i == 2:
                cell.fill = PatternFill("solid", fgColor=GOLD)
                cell.font = Font(bold=True)

    wb.save(filepath)
    print(f"[DCF] Formula-driven model saved to: {filepath}")
    return filepath